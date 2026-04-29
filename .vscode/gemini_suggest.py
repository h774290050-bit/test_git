import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# ==================== 辅助工具函数 ====================
def is_numeric_string(val):
    if not isinstance(val, str): return False
    s = val.strip().replace(',', '')
    if not s: return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def convert_to_number(val):
    if not isinstance(val, str): return val
    s = val.strip().replace(',', '')
    try:
        num = float(s_clean := s)
        return int(num) if num.is_integer() else num
    except ValueError:
        return val


# ==================== 样式与表头逻辑 (已优化传参) ====================
def create_sales_order_header(ws):
    """
    完全基于原代码逻辑修正版：
    1. 确保 BJ-BL, BP-BR, BX-CC 严格执行 4-5 行纵向合并
    2. 确保 A-AH 严格执行 4-5 行纵向合并
    3. AI-BI, BM-BO, BS-BW 保持横向大标题+第5行子标题模式
    """
    ws.title = "销售订单"

    # --- 1. 样式定义 (保持原代码风格) ---
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    header_font = Font(name='微软雅黑', size=10, bold=True)
    big_header_font = Font(name='宋体', size=14, bold=True)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- 2. 第一行：说明文本 (A1:CE1) ---
    content_text = (
        "说明：\n"
        "1、应收金额=商品金额-联合营销费-限时折扣/定金预售报名价\n"
        "2、联合营销费+商家返利金额=0。其中商家返利金额抵扣服务费，开票时会减掉该金额；\n"
        "3、合计技术服务费=技术服务费-商家返利+技术服务费券+技术服务费差异调整；\n"
        "4、转账手续费=订单实际支付成功金额*1%\n"
        "5、分期免息卖家承担金额：商家报名免息活动后订单成交后，商家需要承担的成本；\n"
        "6、调整金额=技术服务费差异调整+转账手续费差异调整+操作费调整+卖家承担邮费调整+卖家承担折扣活动金额调整；\n"
        "7、分销服务费：为按照您设置的比例，应分给导购人的佣金。"
    )
    ws["A1"] = content_text
    ws.merge_cells(f"A1:{get_column_letter(83)}1")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
    ws["A1"].font = Font(name='微软雅黑', size=10)
    ws["A1"].border = thin_border
    ws.row_dimensions[1].height = 140

    # --- 3. 第二行：SUM公式 (A2:CE2) ---
    for col_idx in range(1, 84):
        col_l = get_column_letter(col_idx)
        ws.cell(row=2, column=col_idx, value=f"=SUM({col_l}6:{col_l}30000)")

    # --- 4. 第三行：四大块标题 ---
    big_blocks = [("A", "O", "订单基础信息"), ("P", "AH", "平台服务费信息（一口价）"),
                  ("AI", "BP", "平台服务费信息"), ("BQ", "CC", "结算信息")]
    for s, e, txt in big_blocks:
        ws.merge_cells(f"{s}3:{e}3")
        cell = ws[f"{s}3"]
        cell.value = txt
        cell.alignment = center_alignment
        cell.font = big_header_font
        cell.border = thin_border
    ws.row_dimensions[3].height = 30

    # --- 5. 核心逻辑：【纵向合并区】第4-5行合并 (A-AH, BJ-BL, BP-BR, BX-CC) ---

    # A-AH 的标题
    headers_A_to_AH = [
        "订单号", "订单类型", "商品名称", "商品货号", "数量", "规格", "预约单号", "商品金额",
        "联合营销费", "限时折扣/定金预售报名价", "商品交易金额", "出价时间", "订单创建时间",
        "支付时间", "发货时间", "是否参加活动", "活动费率", "费率活动ID", "适用费率",
        "费率下限", "费率上限", "优惠①:费率折扣", "服务分费率折扣", "费率折扣优惠额",
        "任务达成折扣", "任务达成折扣对应优惠额", "其中-服务费返利折扣金额", "优惠②.技术服务费券",
        "优惠③服务费返利减免金额", "合计平台服务费金额(已扣减优惠①②③)", "商家返利",
        "最终平台基础服务费", "其中:基础服务费金额", "其中:履约服务费金额"
    ]

    # 其他纵向合并的区域 (对应你原代码的 merge_regions)
    vertical_merge_configs = [
        (1, headers_A_to_AH),  # A列开始
        (column_index_from_string("BJ"), ["售后无忧服务费", "出口推广服务费", "卖家退运服务费"]),
        (column_index_from_string("BP"), ["合计平台服务费", "平台预付款收回金额", "以旧换新补贴金额"]),
        (column_index_from_string("BX"),
         ["售中降价(退款)", "售中降价(退津贴)", "调整金额", "应结金额", "结算状态", "结算渠道"])
    ]

    for start_idx, header_list in vertical_merge_configs:
        for offset, text in enumerate(header_list):
            col_l = get_column_letter(start_idx + offset)
            ws.merge_cells(f"{col_l}4:{col_l}5")  # 严格纵向合并
            cell = ws[f"{col_l}4"]
            cell.value = text
            cell.alignment = center_alignment
            cell.font = header_font
            cell.border = thin_border

    # --- 6. 核心逻辑：【横向大标题区】第4行合并 + 第5行填子标题 (AI-BI, BM-BO, BS-BW, CD-CE) ---

    # 对应你原代码的 fill_row5 逻辑
    sub_header_configs = [
        ("AI", "AK", "技术服务费活动信息", ["是否参加活动", "活动费率", "费率活动ID"]),
        ("AL", "AY", "技术服务费信息",
         ["技术服务费费率", "费率下限", "费率上限", "优惠①:费率折扣", "服务分费率折扣", "费率折扣优惠额",
          "任务达成折扣", "任务达成折扣对应优惠额", "其中-服务费返利折扣金额", "优惠②.技术服务费券",
          "优惠③服务费返利减免金额", "技术服务费(已扣减优惠①②③)", "商家返利", "合计技术服务费"]),
        ("AZ", "BC", "操作服务费信息", ["操作服务费", "包含防尘袋包装费", "包含礼盒费", "包含礼袋费"]),
        ("BD", "BI", "操作类费用", ["查验费", "鉴别费", "包装服务费", "转账手续费", "品牌服务费", "客服托管服务费"]),
        ("BM", "BO", "分销服务费", ["分销服务费金额", "分销规则类型", "分销规则ID"]),
        ("BS", "BW", "卖家补贴金额",
         ["卖家承担包邮金额", "消费者邮费补贴金额", "卖家承担优惠券金额", "卖家承担折扣活动金额",
          "分期免息卖家承担金额"]),
        ("CD", "CE", "结算明细", ["货款", "数量"])
    ]

    for start_col, end_col, big_title, sub_list in sub_header_configs:
        # 第4行横向合并
        ws.merge_cells(f"{start_col}4:{end_col}4")
        ws[f"{start_col}4"] = big_title
        ws[f"{start_col}4"].fill = gray_fill
        ws[f"{start_col}4"].alignment = center_alignment
        ws[f"{start_col}4"].font = header_font

        # 第5行填充子标题
        start_idx = column_index_from_string(start_col)
        for offset, sub_txt in enumerate(sub_list):
            col_l = get_column_letter(start_idx + offset)
            cell = ws[f"{col_l}5"]
            cell.value = sub_txt
            cell.alignment = center_alignment
            cell.font = header_font
            cell.border = thin_border

    # --- 7. 善后：全表样式刷 (3-5行) ---
    for row in range(3, 6):
        for col in range(1, 84):
            cell = ws.cell(row=row, column=col)
            cell.fill = gray_fill
            cell.border = thin_border
            if row == 3:
                cell.font = big_header_font
            else:
                cell.font = header_font

    # --- 8. 公式填充 ---
    for row in range(6, 30001):
        ws[f"CD{row}"] = f"=K{row}+BU{row}+BV{row}+BX{row}+BY{row}+BS{row}"

    ws.row_dimensions[4].height = 40
    return ws


def create_Return_order_header(wb):
    """
    重构后的退货退款订单表头创建函数
    """
    # 1. 创建并获取工作表
    ws = wb.create_sheet("退货退款订单")

    # --- 定义通用样式 ---
    style_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    style_alignment_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
    style_font_bold = Font(name='微软雅黑', size=10, bold=True)
    style_font_header = Font(name='宋体', size=14, bold=True)
    style_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill_grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # --- 2. 第一行：说明文本 ---
    content_text_return = (
        "说明：\n"
        "1、应收金额=商品金额-联合营销费-限时折扣/定金预售报名价\n"
        "2、联合营销费+商家返利金额=0。其中商家返利金额抵扣服务费，开票时会减掉该金额；\n"
        "3、合计技术服务费=技术服务费-商家返利+技术服务费券+技术服务费差异调整；\n"
        "4、转账手续费=订单实际支付成功金额*1%\n"
        "5、分期免息卖家承担金额：商家报名免息活动后订单成交后，商家需要承担的成本；\n"
        "6、调整金额=技术服务费差异调整+转账手续费差异调整+操作费调整+卖家承担邮费调整+卖家承担折扣活动金额调整；\n"
        "7、分销服务费：为按照您设置的比例，应分给导购人的佣金。"
    )
    ws["A1"] = content_text_return
    ws.merge_cells(f"A1:{get_column_letter(83)}1") # 合并到CE列附近
    ws["A1"].alignment = style_alignment_top_left
    ws["A1"].font = Font(name='微软雅黑', size=10)
    ws.row_dimensions[1].height = 140

    # --- 3. 第二行：SUM公式区域 (A-CE列) ---
    for col_idx in range(1, 84): # 1到83列 (A到CE)
        col_letter = get_column_letter(col_idx)
        ws.cell(row=2, column=col_idx, value=f"=SUM({col_letter}6:{col_letter}30000)")

    # --- 4. 第三行：大块标题 ---
    big_headers = [
        ("A", "R", "订单基础信息"),
        ("S", "AK", "平台服务费信息（一口价）"),
        ("AL", "BN", "平台服务费信息"),
        ("BO", "BZ", "结算信息"),
        ("CA", "CA", "退货信息") # 单列也统一处理样式
    ]
    for start, end, text in big_headers:
        if start != end:
            ws.merge_cells(f"{start}3:{end}3")
        cell = ws[f"{start}3"]
        cell.value = text
        cell.alignment = style_alignment_center
        cell.font = style_font_header
        cell.fill = fill_grey
        cell.border = style_border
    ws.row_dimensions[3].height = 30

    # --- 5. 第四、五行：复杂表头配置 ---

    # A-AK 列：纵向合并 4-5 行
    headers_A_to_AK = [
        "订单号", "退货订单号", "退货创建时间", "退货订单账单起止时间", "订单类型",
        "商品名称", "商品货号", "数量", "规格", "预约单号",
        "商品金额", "联合营销费", "限时折扣/定金预售报名价", "商品交易金额", "出价时间",
        "订单创建时间", "支付时间", "发货时间", "是否参加活动", "活动费率","费率活动ID",
        "适用费率","费率下限","费率上限","优惠①:费率折扣","服务分费率折扣","费率折扣优惠额","任务达成折扣",
        "任务达成折扣对应优惠额","其中-服务费返利折扣金额","优惠②.技术服务费券",
        "优惠③服务费返利减免金额","合计平台服务费金额(已扣减优惠①②③)","商家返利",
        "最终平台基础服务费", "其中:基础服务费金额", "其中:履约服务费金额"
    ]
    for i, text in enumerate(headers_A_to_AK, 1):
        col_l = get_column_letter(i)
        ws.merge_cells(f"{col_l}4:{col_l}5")
        cell = ws[f"{col_l}4"]
        cell.value = text
        cell.alignment = style_alignment_center
        cell.font = style_font_bold
        cell.border = style_border

    # AL-BU 区域：横向合并 + 子列明细
    sub_configs = [
        ("AL", "AN", "技术服务费活动信息", ["是否参加活动", "活动费率", "费率活动ID"]),
        ("AO", "BB", "技术服务费信息", [
            "技术服务费费率", "费率下限", "费率上限", "优惠①:费率折扣", "服务分费率折扣",
            "费率折扣优惠额", "任务达成折扣", "任务达成折扣对应优惠额", "其中-服务费返利折扣金额",
            "优惠②.技术服务费券", "优惠③服务费返利减免金额", "技术服务费(已扣减优惠①②③)",
            "商家返利", "合计技术服务费"
        ]),
        ("BC", "BF", "操作服务费信息", ["操作服务费", "包含防尘袋包装费", "包含礼盒费", "包含礼袋费"]),
        ("BG", "BK", "操作类费用", ["查验费", "鉴别费", "包装服务费", "转账手续费", "客服托管服务费"]),
        ("BL", "BL", "其他费用", ["品牌服务费"]),
        ("BQ", "BU", "卖家补贴金额", [
            "卖家承担包邮金额", "消费者邮费补贴金额", "卖家承担优惠券金额", "卖家承担折扣活动金额", "分期免息卖家承担金额"
        ])
    ]

    for start_col, end_col, top_text, sub_headers in sub_configs:
        if start_col != end_col:
            ws.merge_cells(f"{start_col}4:{end_col}4")
        ws[f"{start_col}4"] = top_text
        ws[f"{start_col}4"].alignment = style_alignment_center
        ws[f"{start_col}4"].font = style_font_bold
        ws[f"{start_col}4"].fill = fill_grey

        start_idx = column_index_from_string(start_col)
        for offset, text in enumerate(sub_headers):
            cell = ws.cell(row=5, column=start_idx + offset, value=text)
            cell.alignment = style_alignment_center
            cell.font = style_font_bold
            cell.border = style_border

    # BM-BP & BV-CB：纵向合并 4-5 行
    extra_merge_configs = [
        (65, ["合计平台服务费", "售后无忧服务费", "平台预付款收回金额", "以旧换新补贴金额"]), # 从BM(65)开始
        (74, ["售中降价(退款)", "售中降价(退津贴)", "调整金额", "应结金额", "结算状态", "收款账期", "货款"]) # 从BV(74)开始
    ]
    for start_idx, header_list in extra_merge_configs:
        for offset, text in enumerate(header_list):
            col_l = get_column_letter(start_idx + offset)
            ws.merge_cells(f"{col_l}4:{col_l}5")
            ws[f"{col_l}4"] = text
            ws[f"{col_l}4"].alignment = style_alignment_center
            ws[f"{col_l}4"].font = style_font_bold
            ws[f"{col_l}4"].border = style_border

    # 特殊单格处理
    ws["CC5"] = "数量"
    ws["CC5"].font = style_font_bold
    ws["CC5"].alignment = style_alignment_center

    # --- 6. 公式与整体样式修饰 ---
    # CB列计算公式
    for row in range(6, 30001):
        ws[f"CB{row}"] = f"=N{row}+BS{row}+BT{row}+BV{row}+BQ{row}"

    # 统一填充灰色和边框 (3-5行)
    for r in range(3, 6):
        for c in range(1, 84):
            cell = ws.cell(row=r, column=c)
            cell.border = style_border
            if not cell.fill.start_color.index:
                cell.fill = fill_grey

    ws.row_dimensions[4].height = 40
    return ws


def create_full_audit_excel(wb, year_month, shop_code):
    """
    重构后的稽核表创建函数
    """
    ws = wb.create_sheet("稽核")

    # --- 1. 定义样式 ---
    style_font_bold = Font(name='微软雅黑', size=10, bold=True)
    style_alignment_center = Alignment(horizontal='center', vertical='center')
    style_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill_header = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # --- 2. 顶部信息 (第1-2行) ---
    ws['A1'] = '年月'
    ws['B1'] = year_month
    ws['E1'] = '新模版-检查公式'
    ws['A1'].font = style_font_bold

    # --- 3. 写入主表头 (第3行) ---
    headers = [
        "账单类别", "年月", "erp店号", "账户主体", "收入金额（+元）", "支出金额（-元）",
        "一级分类", "二级分类", "科目代码", "科目名称", "对账年月", "结算账号",
        "供应商", "资金账户", "应收余额", "合计"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = style_font_bold
        cell.alignment = style_alignment_center
        cell.fill = fill_header
        cell.border = style_border

    # --- 4. 填充固定数据 (第4-28行) ---
    # 定义左侧四列的规律数据
    row_configs = [
        ("账期对账单", 8), # 类别名称, 重复次数
        ("支付宝账单", 5),
        ("银行流水", 12)
    ]

    current_row = 4
    for label, repeat in row_configs:
        for _ in range(repeat):
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=year_month)
            ws.cell(row=current_row, column=3, value="DW01")
            ws.cell(row=current_row, column=4, value="HN")
            current_row += 1

    # --- 5. 写入核心公式 (第五、六列) ---
    # E4 单元格的长公式
    ws['E4'] = (
        "=SUM('销售订单'!K2,'销售订单'!BU2,'销售订单'!BV2,'销售订单'!BX2,'销售订单'!BY2,'销售订单'!BS2,"
        "'退货退款订单'!N2,'退货退款订单'!BS2,'退货退款订单'!BT2,'退货退款订单'!BV2,'退货退款订单'!BW2,'退货退款订单'!BQ2)"
    )

    # E17:E28 的银行收款记录公式
    for i, row_idx in enumerate(range(17, 29)):
        ws.cell(row=row_idx, column=5, value=f"=银行收款记录!P{i+2}")

    # F4:F11 的支出/费用公式
    expenditure_formulas = [
        "=(销售订单!L2+销售订单!BE2)+(退货退款订单!O2+退货退款订单!AV2)",
        "=销售订单!AY2+退货退款订单!BB2",
        "=销售订单!BG2+退货退款订单!BJ2",
        "=销售订单!BH2+退货退款订单!BL2",
        "=销售订单!BT2+退货退款订单!BR2",
        "=销售订单!BI2+退货退款订单!BK2",
        "=SUM(D34:D40)", # 引用外部
        "=销售订单!AF2+退货退款订单!AI2"
    ]
    for i, formula in enumerate(expenditure_formulas):
        ws.cell(row=4 + i, column=6, value=formula)

    # --- 6. 填充第七至十四列 (财务科目信息) ---
    finance_info = [
        ["回款", "货款", "1122.002.02.1300", "应收账款-第三方-自营&电商&代销-A/R trade", "", "", "", ""],
        ["费用", "技术服务费", "6601.006.98.8133", "期间费用-电商费用-其他-Outside services", "", "", "上海识装信息科技有限公司", ""],
        ["费用", "转账手续费", "6601.029.99.8133", "期间费用-支付服务费--Outside services", "", "", "上海识装信息科技有限公司", ""],
        ["费用", "品牌服务费", "6601.029.99.8133", "期间费用-电商费用-其他-Outside services", "", "", "上海识装信息科技有限公司", ""],
        ["费用", "消费者邮费", "6601.006.98.8133", "期间费用-电商费用-其他-Outside services", "", "", "上海识装信息科技有限公司", ""],
        ["费用", "客服托管服务费", "", "", "", "", "", ""],
        ["回款", "交易售后赔付款", "1122.002.02.1300", "应收账款-第三方-自营&电商&代销-A/R trade", "", "", "", ""],
        ["回款", "平台基础服务费（一口价）", "6601.006.98.8133", "期间费用-电商费用-其他-Outside services", "", "", "", ""],
        # 后面是重复的“平台拨款”
        *( [["回款", "平台拨款", "", "", "", "", "", "支付宝"]] * 5 ),
        *( [["回款", "平台拨款", "", "", "", "", "", "中信银行"]] * 12 )
    ]
    for i, row_data in enumerate(finance_info):
        for j, value in enumerate(row_data):
            ws.cell(row=4 + i, column=7 + j, value=value)

    # --- 7. 第十五、十六列：余额与合计公式 ---
    # O4 初始余额
    ws['O4'] = "=E4+F4"
    # O5:O11 累加
    for i in range(5, 12):
        ws.cell(row=i, column=15, value=f"=O{i-1}+E{i}+F{i}")
    # O12:O28 扣减
    for i in range(12, 29):
        ws.cell(row=i, column=15, value=f"=O{i-1}-E{i}-F{i}")
    # P列 每行合计
    for i in range(4, 29):
        ws.cell(row=i, column=16, value=f"=E{i}+F{i}")

    # --- 8. 底部“扣减其他费用”区域 (修正位置与合并逻辑) ---
    footer_row = 32  # 标题回到第32行
    ws.cell(row=footer_row, column=1, value="扣减其他费用:交易售后赔付款").font = style_font_bold

    # 表头行：需求是“余下表头从33行开始”
    f_headers = ["费用类型", "偿还总金额", "明细费用项", "偿还金额", "币种"]
    header_row_idx = 33  # 明确指定为33行
    for col, text in enumerate(f_headers, 1):
        cell = ws.cell(row=header_row_idx, column=col, value=text)
        cell.font = style_font_bold
        cell.border = style_border
        cell.alignment = style_alignment_center

    # 数据行：从34行开始
    f_details = [
        "价保差额补偿款",
        "交易售后赔付款",
        "交易售后赔付款(其他)"
    ]

    start_data_row = 34
    for i, detail_text in enumerate(f_details):
        curr_row = start_data_row + i
        # 填充边框和对齐样式
        for col in range(1, 6):
            ws.cell(row=curr_row, column=col).border = style_border
            ws.cell(row=curr_row, column=col).alignment = style_alignment_center

        # C列：明细项名称
        ws.cell(row=curr_row, column=3, value=detail_text)
        # E列：币种
        ws.cell(row=curr_row, column=5, value="CNY")

    # --- 执行合并需求 ---
    # 1. 合并 A34:A36 (费用类型)，填入“服务费”
    ws.merge_cells(start_row=start_data_row, start_column=1, end_row=start_data_row + 2, end_column=1)
    ws.cell(row=start_data_row, column=1, value="服务费")

    # 2. 合并 B34:B36 (偿还总金额)，不填数据
    ws.merge_cells(start_row=start_data_row, start_column=2, end_row=start_data_row + 2, end_column=2)

    # D列（偿还金额）已经在循环中通过 border 逻辑绘制了边框且未填值，自动留空

    # --- 9. Q列 辅助校验公式 ---
    ws['Q4'] = "=E4+F8+F11+F10-SUM(E12:E28)"
    ws['Q10'] = "=E10-P10"

    return ws


def create_bank_receipt_records(wb):
    """创建银行记录表"""
    ws4 = wb.create_sheet("银行收款记录")
    return ws4


# ==================== 核心数据搬运逻辑 ====================


# ==================== 流程控制逻辑 ====================

def copy_data_smart_convert(src_file, dst_wb, sheet_name):
    """
    修正版：智能搬运数据并进行数值化转换
    """
    print(f"  🚀 正在处理 [{sheet_name}] 数据搬运...")

    # 1. 加载源工作簿 (data_only=True 确保读取的是计算后的值而不是公式)
    try:
        src_wb = load_workbook(src_file, data_only=True)
    except Exception as e:
        print(f"  ❌ 无法打开源文件: {e}")
        return

    if sheet_name not in src_wb.sheetnames:
        print(f"  ⚠️ 源文件中不存在工作表: {sheet_name}")
        return

    src_ws = src_wb[sheet_name]
    dst_ws = dst_wb[sheet_name]

    # === 配置区域 ===
    SRC_HEADER_ROW = 4  # 源文件表头所在行
    SRC_DATA_START = 5  # 源文件数据起始行 (强制从第5行开始读取)
    DST_HEADER_ROW = 5  # 模板文件表头所在行
    DATA_START_ROW = 6  # 模板文件数据写入起始行

    # 2. 获取表头（处理合并单元格逻辑）
    def get_smart_headers(ws, target_row):
        headers = []
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            val = ws.cell(row=target_row, column=col).value
            # 如果目标行单元格为空，尝试向上找一行（应对合并单元格）
            if val is None or str(val).strip() == "":
                val = ws.cell(row=target_row - 1, column=col).value
            headers.append(str(val).strip() if val is not None else "")
        return headers

    src_headers = get_smart_headers(src_ws, SRC_HEADER_ROW)
    dst_headers = get_smart_headers(dst_ws, DST_HEADER_ROW)

    # 3. 建立列名映射关系
    col_mapping = []
    for dst_h in dst_headers:
        if dst_h != "" and dst_h in src_headers:
            col_mapping.append(src_headers.index(dst_h))
        else:
            col_mapping.append(None)

    # 4. 循环读取并写入数据
    write_count = 0  # 实际写入的数据行数
    converted_count = 0  # 发生数值转换的次数

    # 使用 iter_rows 从第6行开始读取
    for row_data in src_ws.iter_rows(min_row=SRC_DATA_START, values_only=True):

        # 严格空行判定：如果这一行所有选定的映射列都是空的，则跳过
        # 这样比 all(v is None for v in row_data) 更稳健
        has_data = False
        for src_idx in col_mapping:
            if src_idx is not None and row_data[src_idx] is not None:
                if str(row_data[src_idx]).strip() != "":
                    has_data = True
                    break

        if not has_data:
            continue

        # 计算目标行号：起始行(6) + 当前已写入数量
        current_target_row = DATA_START_ROW + write_count

        for dst_col_idx, src_idx in enumerate(col_mapping):
            if src_idx is not None:
                # 获取源数据
                original_val = row_data[src_idx]

                # 智能处理数值转换
                final_val = original_val
                if original_val is not None:
                    if is_numeric_string(original_val):
                        final_val = convert_to_number(original_val)
                        converted_count += 1

                # 写入目标 Sheet
                dst_ws.cell(row=current_target_row, column=dst_col_idx + 1, value=final_val)

        write_count += 1

        if write_count % 500 == 0:
            print(f"    ...已搬运 {write_count} 行...")

    print(f"  ✅ [{sheet_name}] 处理完毕：共搬运 {write_count} 行，数值化转换 {converted_count} 处。")

def process_single_file(file_path, output_path):
    """
    处理单个文件的完整生命周期
    """
    # 1. 准备动态参数
    # 假设文件名是 "DW01 账单...", 截取前4位即 "DW01"
    shop_code = file_path.name[:4]
    
    # 1. 创建新工作簿并初始化结构
    wb = Workbook()
    ws_sales = wb.active

    # 2. 调用你定义的复杂表头函数
    create_sales_order_header(ws_sales)
    create_Return_order_header(wb)
    create_bank_receipt_records(wb)
    ws_audit = create_full_audit_excel(wb, year_month="202603",shop_code=shop_code)

    # 调整 Sheet 顺序（让稽核表排在第一位）
    sheets = wb._sheets
    sheets.insert(0, sheets.pop(sheets.index(ws_audit)))

    # 3. 数据转换与填充
    copy_data_smart_convert(file_path, wb, "销售订单")
    copy_data_smart_convert(file_path, wb, "退货退款订单")

    # 4. 保存到新路径
    wb.save(output_path)


def batch_process(folder_path):
    base_dir = Path(folder_path)
    target_dir = base_dir / "分类"
    target_dir.mkdir(exist_ok=True)

    # 排除掉已经是处理结果的文件
    files = [f for f in base_dir.glob("*.xlsx") if "分类" not in f.parts]

    for file_path in files:
        print(f"🔎 正在处理: {file_path.name}")
        output_path = target_dir / f"已处理_{file_path.name}"

        try:
            process_single_file(file_path, output_path)
            print(f"✅ 处理成功 -> {output_path}")
        except Exception as e:
            print(f"❌ 处理失败 {file_path.name}: {e}")


if __name__ == "__main__":
    path = input("如若需要更改年月，在第570行更改；请输入文件夹路径: ").strip('"')
    if os.path.exists(path):
        batch_process(path)
    else:
        print("路径不存在！")

#4月28日定稿封存